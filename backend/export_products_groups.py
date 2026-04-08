"""
Экспорт уникальных пар КОД ТОВАРА + ГРУППА ТОВАРА из таблицы sales в Excel.

Использование:
    python export_products_groups.py [--output имя_файла.xlsx]
"""

import os
import sys
import argparse
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from api.models import Sale


def export(output_path: str):
    print("Читаю уникальные значения из таблицы sales...")

    rows = (
        Sale.objects
        .exclude(kod_tovara='')
        .filter(kod_tovara__isnull=False)
        .values('kod_tovara', 'gruppa_tovara')
        .distinct()
        .order_by('gruppa_tovara', 'kod_tovara')
    )

    data = list(rows)
    print(f"Найдено уникальных пар: {len(data)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Коды и группы"

    # Заголовки
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["КОД ТОВАРА", "ГРУППА ТОВАРА"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные
    for row_idx, item in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=item['kod_tovara'] or '')
        ws.cell(row=row_idx, column=2, value=item['gruppa_tovara'] or '')
        # Чередование строк
        if row_idx % 2 == 0:
            fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            for col in range(1, 3):
                ws.cell(row=row_idx, column=col).fill = fill

    # Ширина колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35

    # Заморозить шапку
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Файл сохранён: {output_path}")
    print(f"Итого строк: {len(data)}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--output', default='Коды и группы товаров.xlsx',
                        help='Имя выходного файла (по умолчанию: "Коды и группы товаров.xlsx")')
    args = parser.parse_args()
    export(args.output)
