"""
Генерация Excel шаблона для загрузки данных.
Берёт реальные строки из базы и вставляет в шаблон как примеры.
"""
import django, os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from api.models import Sale, ReadySale
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── Стили ────────────────────────────────────────────────────────────────────
HEADER_FILL     = PatternFill("solid", fgColor="1F3864")   # тёмно-синий
REQUIRED_FILL   = PatternFill("solid", fgColor="C00000")   # красный
EXAMPLE_FILL    = PatternFill("solid", fgColor="E8F0FE")   # голубоватый
HEADER_FONT     = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
REQUIRED_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT       = Font(name="Calibri", size=10)
EXAMPLE_FONT    = Font(name="Calibri", size=10, color="444444", italic=True)
THIN_BORDER     = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")


def style_header(cell, required=False):
    cell.font      = REQUIRED_FONT if required else HEADER_FONT
    cell.fill      = REQUIRED_FILL if required else HEADER_FILL
    cell.alignment = CENTER
    cell.border    = THIN_BORDER


def style_example(cell):
    cell.font      = EXAMPLE_FONT
    cell.fill      = EXAMPLE_FILL
    cell.alignment = LEFT
    cell.border    = THIN_BORDER


def style_cell(cell):
    cell.font      = CELL_FONT
    cell.alignment = LEFT
    cell.border    = THIN_BORDER


def set_column_widths(ws, widths):
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ─── Реальные данные из базы ──────────────────────────────────────────────────
SAMPLE_SALES = []
for s in Sale.objects.exclude(kod_tovara='').exclude(data=None)[:7]:
    SAMPLE_SALES.append([
        s.kod_tovara,
        s.gruppa_tovara,
        s.region,
        s.sklad,
        s.scheta,
        str(s.data)[:10] if s.data else '',
        s.dopoln_kol_vo,
        s.uch_kol_vo,
        s.tovary,
        s.cvet,
        s.profil_perechen,
    ])

SAMPLE_READY = []
for r in ReadySale.objects.exclude(klient=None).exclude(data=None)[:7]:
    SAMPLE_READY.append([
        str(r.data) if r.data else '',
        r.klient,
        r.diler,
        r.region,
        r.klient_id,
        r.invoice_sid,
        r.tip,
        r.tip_organizacii,
        r.produkt,
        r.gruppa_produktov,
        r.kolichestvo,
        r.ves_kg,
        r.obshchaya_summa,
        r.dokhod,
        r.valyuta,
        r.tovary,
    ])


ws1 = wb.active
ws1.title = "Sales"

# Строка 1 — заголовок раздела
ws1.merge_cells("A1:J1")
title_cell = ws1["A1"]
title_cell.value = "ШАБЛОН ЗАГРУЗКИ: Sales  (data_type = sales)"
title_cell.font  = Font(name="Calibri", bold=True, size=13, color="1F3864")
title_cell.alignment = CENTER
ws1.row_dimensions[1].height = 28

# Строка 2 — подсказка
ws1.merge_cells("A2:J2")
hint = ws1["A2"]
hint.value = ("Красный заголовок = обязательное поле.  "
              "Дата формат: YYYY-MM-DD  |  ДОПОЛН__КОЛ-ВО — вес (кг), УЧИТЫВАЯ_КОЛ-ВО — штуки")
hint.font  = Font(name="Calibri", size=9, italic=True, color="666666")
hint.alignment = LEFT
ws1.row_dimensions[2].height = 18

# Строка 3 — заголовки колонок
SALES_COLS = [
    ("КОД_ТОВАРА",        True),
    ("ГРУППА_ТОВАРА",     False),
    ("РЕГИОН",            False),
    ("СКЛАД",             False),
    ("СЧЕТЫ",             False),
    ("Дата",              False),
    ("ДОПОЛН__КОЛ-ВО",   False),
    ("УЧИТЫВАЯ_КОЛ-ВО",  False),
    ("ТОВАРЫ",            False),
    ("ЦВЕТ",              False),
    ("профиль_перечень",  False),
]

ws1.row_dimensions[3].height = 36
for col_idx, (col_name, required) in enumerate(SALES_COLS, start=1):
    cell = ws1.cell(row=3, column=col_idx, value=col_name)
    style_header(cell, required)

# Строки с реальными данными
for row_num, row_data in enumerate(SAMPLE_SALES, start=4):
    fill = EXAMPLE_FILL if row_num == 4 else PatternFill("solid", fgColor="F8F9FF")
    ws1.row_dimensions[row_num].height = 18
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_num, column=col_idx, value=value)
        cell.font = EXAMPLE_FONT
        cell.fill = fill
        cell.alignment = LEFT
        cell.border = THIN_BORDER

# Пустые строки после примеров
for row in range(4 + len(SAMPLE_SALES), 4 + len(SAMPLE_SALES) + 5):
    ws1.row_dimensions[row].height = 18
    for col_idx in range(1, len(SALES_COLS) + 1):
        style_cell(ws1.cell(row=row, column=col_idx))

set_column_widths(ws1, [16, 16, 14, 10, 10, 14, 18, 18, 24, 14, 18])

# Заморозить первые 3 строки
ws1.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════════
# Лист 2 — ReadySales (тип: ready_sales)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("ReadySales")

ws2.merge_cells("A1:P1")
title2 = ws2["A1"]
title2.value = "ШАБЛОН ЗАГРУЗКИ: ReadySales  (data_type = ready_sales)"
title2.font  = Font(name="Calibri", bold=True, size=13, color="1F3864")
title2.alignment = CENTER
ws2.row_dimensions[1].height = 28

ws2.merge_cells("A2:P2")
hint2 = ws2["A2"]
hint2.value = ("Красный заголовок = обязательное поле.  "
               "Дата формат: YYYY-MM-DD  |  Вес_кг — числовое, Количество — числовое (шт)")
hint2.font  = Font(name="Calibri", size=9, italic=True, color="666666")
hint2.alignment = LEFT
ws2.row_dimensions[2].height = 18

READY_COLS = [
    ("Дата",              True),
    ("Клиент",            True),
    ("Дилер",             False),
    ("Регион",            False),
    ("Клиент_ИД",         False),
    ("Инвоиcе_CИД",       False),
    ("Тип",               False),
    ("Тип_организации",   False),
    ("Продукт",           False),
    ("Группа_продуктов",  False),
    ("Количество",        False),
    ("Вес_кг",            False),
    ("Общая_сумма",       False),
    ("Доход",             False),
    ("Валюта",            False),
    ("ТОВАРЫ",            False),
]

ws2.row_dimensions[3].height = 36
for col_idx, (col_name, required) in enumerate(READY_COLS, start=1):
    cell = ws2.cell(row=3, column=col_idx, value=col_name)
    style_header(cell, required)

# Строки с реальными данными
for row_num, row_data in enumerate(SAMPLE_READY, start=4):
    fill = EXAMPLE_FILL if row_num == 4 else PatternFill("solid", fgColor="F8F9FF")
    ws2.row_dimensions[row_num].height = 18
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_num, column=col_idx, value=value)
        cell.font = EXAMPLE_FONT
        cell.fill = fill
        cell.alignment = LEFT
        cell.border = THIN_BORDER

# Пустые строки после примеров
for row in range(4 + len(SAMPLE_READY), 4 + len(SAMPLE_READY) + 5):
    ws2.row_dimensions[row].height = 18
    for col_idx in range(1, len(READY_COLS) + 1):
        style_cell(ws2.cell(row=row, column=col_idx))

set_column_widths(ws2, [14, 24, 20, 14, 11, 14, 12, 20, 24, 20, 12, 10, 14, 12, 10, 14])
ws2.freeze_panes = "A4"


# ─── Сохраняем ────────────────────────────────────────────────────────────────
output_path = "upload_template.xlsx"
wb.save(output_path)
print(f"✓ Шаблон сохранён: {output_path}")
print()
print("Листы:")
print("  Sales      → data_type=sales      (обязательно: КОД_ТОВАРА)")
print("  ReadySales → data_type=ready_sales (обязательно: Дата, Клиент)")
